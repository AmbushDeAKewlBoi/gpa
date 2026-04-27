from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation


ROOT = Path(__file__).resolve().parents[1]
OUTPUT = ROOT / "LockIn_GPA_Calculator.xlsx"


HEADER_FILL = PatternFill("solid", fgColor="0F172A")
SECTION_FILL = PatternFill("solid", fgColor="1E293B")
ACCENT_FILL = PatternFill("solid", fgColor="0B3B5B")
TEXT_LIGHT = "E2E8F0"
TEXT_MUTED = "94A3B8"
BORDER = Border(
    left=Side(style="thin", color="334155"),
    right=Side(style="thin", color="334155"),
    top=Side(style="thin", color="334155"),
    bottom=Side(style="thin", color="334155"),
)


def style_header(cell):
    cell.fill = HEADER_FILL
    cell.font = Font(color=TEXT_LIGHT, bold=True)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = BORDER


def style_section(cell):
    cell.fill = SECTION_FILL
    cell.font = Font(color=TEXT_LIGHT, bold=True)
    cell.border = BORDER


def style_note(cell):
    cell.fill = ACCENT_FILL
    cell.font = Font(color=TEXT_LIGHT, italic=True)
    cell.alignment = Alignment(wrap_text=True, vertical="top")
    cell.border = BORDER


def style_body(cell):
    cell.border = BORDER
    cell.font = Font(color="0F172A")
    cell.alignment = Alignment(vertical="center")


def autosize(ws, widths):
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def add_grade_scale(ws):
    ws.title = "GradeScale"
    headers = ["Min %", "Letter", "Base Points"]
    for idx, label in enumerate(headers, start=1):
        style_header(ws.cell(row=1, column=idx, value=label))

    rows = [
        (97, "A+", 4.3),
        (93, "A", 4.0),
        (90, "A-", 3.7),
        (87, "B+", 3.3),
        (83, "B", 3.0),
        (80, "B-", 2.7),
        (77, "C+", 2.3),
        (73, "C", 2.0),
        (70, "C-", 1.7),
        (67, "D+", 1.3),
        (63, "D", 1.0),
        (60, "D-", 0.7),
        (0, "F", 0.0),
    ]
    for r, values in enumerate(rows, start=2):
        for c, value in enumerate(values, start=1):
            style_body(ws.cell(row=r, column=c, value=value))

    autosize(ws, {"A": 10, "B": 10, "C": 12})
    ws.sheet_state = "hidden"


def add_instructions(ws):
    ws.title = "Instructions"
    ws["A1"] = "LockIn GPA Calculator"
    ws["A1"].font = Font(size=18, bold=True, color=TEXT_LIGHT)
    ws["A1"].fill = HEADER_FILL
    ws["A1"].alignment = Alignment(horizontal="left")

    notes = [
        "1. Enter finished or current classes in the Classes sheet.",
        "2. Use numeric grades if you know them. Letter grades and GPA points fill in automatically.",
        "3. Choose Standard, Honors, AP, or DE in the Level column.",
        "4. Use the Scenario sheet to test future semesters without changing your real classes.",
        "5. Import this workbook into Google Sheets if you want cloud access.",
    ]

    for idx, note in enumerate(notes, start=3):
        style_note(ws.cell(row=idx, column=1, value=note))
        ws.merge_cells(start_row=idx, start_column=1, end_row=idx, end_column=5)

    ws["A10"] = "Weighting used in this workbook"
    style_section(ws["A10"])
    ws["A11"] = "Standard = +0.0"
    ws["A12"] = "Honors = +0.5"
    ws["A13"] = "AP = +1.0"
    ws["A14"] = "DE = +1.0"
    for row in range(11, 15):
        style_body(ws[f"A{row}"])

    autosize(ws, {"A": 60, "B": 16, "C": 16, "D": 16, "E": 16})


def add_course_sheet(ws, title, start_row_count=31):
    ws.title = title
    headers = [
        "Course",
        "Term",
        "Level",
        "Credits",
        "Numeric Grade",
        "Letter Grade",
        "Base Pts",
        "Weight Bonus",
        "Weighted Pts",
        "Unweighted QP",
        "Weighted QP",
    ]
    for idx, label in enumerate(headers, start=1):
        style_header(ws.cell(row=1, column=idx, value=label))

    widths = {
        "A": 28,
        "B": 16,
        "C": 14,
        "D": 10,
        "E": 14,
        "F": 12,
        "G": 10,
        "H": 13,
        "I": 12,
        "J": 14,
        "K": 14,
    }
    autosize(ws, widths)

    level_validation = DataValidation(type="list", formula1='"Standard,Honors,AP,DE"', allow_blank=True)
    term_validation = DataValidation(
        type="list",
        formula1='"Semester 1,Semester 2,Full Year,Summer,Planned"',
        allow_blank=True,
    )
    ws.add_data_validation(level_validation)
    ws.add_data_validation(term_validation)
    level_validation.add(f"C2:C{start_row_count}")
    term_validation.add(f"B2:B{start_row_count}")

    for row in range(2, start_row_count + 1):
        ws[f"D{row}"] = 1
        ws[f"F{row}"] = (
            f'=IF(E{row}="","",LOOKUP(E{row},GradeScale!$A$2:$A$14,GradeScale!$B$2:$B$14))'
        )
        ws[f"G{row}"] = (
            f'=IF(E{row}="","",LOOKUP(E{row},GradeScale!$A$2:$A$14,GradeScale!$C$2:$C$14))'
        )
        ws[f"H{row}"] = (
            f'=IF(C{row}="","",IF(C{row}="Honors",0.5,IF(OR(C{row}="AP",C{row}="DE"),1,0)))'
        )
        ws[f"I{row}"] = f'=IF(G{row}="","",G{row}+H{row})'
        ws[f"J{row}"] = f'=IF(OR(G{row}="",D{row}=""),"",G{row}*D{row})'
        ws[f"K{row}"] = f'=IF(OR(I{row}="",D{row}=""),"",I{row}*D{row})'
        for col in range(1, 12):
            style_body(ws.cell(row=row, column=col))

    ws.freeze_panes = "A2"
    ws.conditional_formatting.add(
        f"E2:E{start_row_count}",
        ColorScaleRule(
            start_type="num",
            start_value=60,
            start_color="F87171",
            mid_type="num",
            mid_value=80,
            mid_color="FACC15",
            end_type="num",
            end_value=100,
            end_color="4ADE80",
        ),
    )


def add_summary(ws):
    ws.title = "Summary"
    ws["A1"] = "GPA Snapshot"
    ws["A1"].font = Font(size=18, bold=True, color=TEXT_LIGHT)
    ws["A1"].fill = HEADER_FILL
    ws["A3"] = "Current Classes"
    style_section(ws["A3"])
    ws["A4"] = "Total Credits"
    ws["A5"] = "Unweighted GPA"
    ws["A6"] = "Weighted GPA"
    ws["A7"] = "Class Count"
    ws["B4"] = '=SUM(Classes!D2:D31)'
    ws["B5"] = '=IFERROR(SUM(Classes!J2:J31)/B4,0)'
    ws["B6"] = '=IFERROR(SUM(Classes!K2:K31)/B4,0)'
    ws["B7"] = '=COUNTA(Classes!A2:A31)'

    ws["D3"] = "Scenario Projection"
    style_section(ws["D3"])
    ws["D4"] = "Added Credits"
    ws["D5"] = "Projected Unweighted GPA"
    ws["D6"] = "Projected Weighted GPA"
    ws["D7"] = "Planned Class Count"
    ws["E4"] = '=SUM(Scenario!D2:D31)'
    ws["E5"] = '=IFERROR((SUM(Classes!J2:J31)+SUM(Scenario!J2:J31))/(B4+E4),0)'
    ws["E6"] = '=IFERROR((SUM(Classes!K2:K31)+SUM(Scenario!K2:K31))/(B4+E4),0)'
    ws["E7"] = '=COUNTA(Scenario!A2:A31)'

    for cell in ("A4", "A5", "A6", "A7", "D4", "D5", "D6", "D7"):
        style_body(ws[cell])
    for cell in ("B4", "B5", "B6", "B7", "E4", "E5", "E6", "E7"):
        style_body(ws[cell])
        ws[cell].number_format = "0.00"
    ws["B4"].number_format = "0.0"
    ws["E4"].number_format = "0.0"
    ws["B7"].number_format = "0"
    ws["E7"].number_format = "0"

    ws["A10"] = "Grade Breakdown"
    style_section(ws["A10"])
    for col, label in zip(("A", "B", "C", "D", "E"), ("A range", "B range", "C range", "D range", "F",)):
        style_body(ws[f"{col}11"])
        ws[f"{col}11"] = label
    ws["A12"] = '=COUNTIF(Classes!E2:E31,">=90")'
    ws["B12"] = '=COUNTIFS(Classes!E2:E31,">=80",Classes!E2:E31,"<90")'
    ws["C12"] = '=COUNTIFS(Classes!E2:E31,">=70",Classes!E2:E31,"<80")'
    ws["D12"] = '=COUNTIFS(Classes!E2:E31,">=60",Classes!E2:E31,"<70")'
    ws["E12"] = '=COUNTIF(Classes!E2:E31,"<60")'
    for col in ("A", "B", "C", "D", "E"):
        style_body(ws[f"{col}12"])

    chart = BarChart()
    chart.title = "Current Grades by Class"
    chart.y_axis.title = "Numeric Grade"
    chart.x_axis.title = "Class"
    chart.height = 7
    chart.width = 12
    values = Reference(ws.parent["Classes"], min_col=5, min_row=1, max_row=31)
    cats = Reference(ws.parent["Classes"], min_col=1, min_row=2, max_row=31)
    chart.add_data(values, titles_from_data=True)
    chart.set_categories(cats)
    ws.add_chart(chart, "G3")

    autosize(ws, {"A": 18, "B": 14, "C": 14, "D": 22, "E": 18, "G": 16})


def main():
    wb = Workbook()
    add_grade_scale(wb.active)
    add_instructions(wb.create_sheet())
    add_course_sheet(wb.create_sheet(), "Classes")
    add_course_sheet(wb.create_sheet(), "Scenario")
    add_summary(wb.create_sheet())

    wb._sheets = [
        wb["Instructions"],
        wb["Classes"],
        wb["Summary"],
        wb["Scenario"],
        wb["GradeScale"],
    ]

    wb.save(OUTPUT)
    print(f"Created {OUTPUT}")


if __name__ == "__main__":
    main()
